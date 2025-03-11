import streamlit as st
import pandas as pd
import calendar
from datetime import datetime, date
import altair as alt
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configuração da página
st.set_page_config(
    page_title="Gerador de Escala de Trabalho",
    page_icon="📅",
    layout="wide"
)

# Função para gerar a escala considerando exceções e equilibrar a carga de trabalho
def gerar_escala(mes, ano, funcionarios, excecoes, considerar_carga=True):
    # Dicionário para controlar a carga de trabalho de cada funcionário
    carga_trabalho = {f: 0 for f in funcionarios}
    
    # Identificar os dias úteis do mês
    dias_uteis = []
    for dia in range(1, calendar.monthrange(ano, mes)[1] + 1):
        data = date(ano, mes, dia)
        dia_semana = data.weekday()
        if dia_semana < 5:  # Segunda (0) a Sexta (4)
            dias_uteis.append(dia)
    
    escala = []
    
    # Para cada dia útil, alocar funcionários aos turnos
    for dia in dias_uteis:
        data = date(ano, mes, dia)
        dia_semana = calendar.day_name[data.weekday()]
        
        # Encontrar funcionários disponíveis para cada turno
        disponiveis_matutino = [f for f in funcionarios if f not in excecoes.get((dia, "Matutino"), [])]
        disponiveis_vespertino = [f for f in funcionarios if f not in excecoes.get((dia, "Vespertino"), [])]
        
        # Se considerar carga de trabalho, ordenar funcionários pelo menor número de turnos alocados
        if considerar_carga:
            disponiveis_matutino.sort(key=lambda f: carga_trabalho[f])
            disponiveis_vespertino.sort(key=lambda f: carga_trabalho[f])
        
        # Alocar funcionários aos turnos
        turno_matutino = disponiveis_matutino[0] if disponiveis_matutino else None
        if turno_matutino:
            carga_trabalho[turno_matutino] += 1
        
        # Tentar não alocar o mesmo funcionário para os dois turnos do mesmo dia
        if turno_matutino in disponiveis_vespertino and len(disponiveis_vespertino) > 1:
            disponiveis_vespertino.remove(turno_matutino)
        
        turno_vespertino = disponiveis_vespertino[0] if disponiveis_vespertino else None
        if turno_vespertino:
            carga_trabalho[turno_vespertino] += 1
        
        escala.append({
            "Data": f"{dia:02d}/{mes:02d}/{ano}",
            "Dia da Semana": dia_semana,
            "Turno Matutino": turno_matutino,
            "Turno Vespertino": turno_vespertino
        })
    
    df = pd.DataFrame(escala)
    
    # Adicionar estatísticas de carga de trabalho
    estatisticas = pd.DataFrame({
        "Funcionário": list(carga_trabalho.keys()),
        "Total de Turnos": list(carga_trabalho.values())
    })
    
    return df, estatisticas

# Função para exportar para Excel com formatação
def exportar_excel(df, estatisticas, mes, ano):
    output = io.BytesIO()
    workbook = Workbook()
    
    # Primeira aba - Escala
    ws_escala = workbook.active
    ws_escala.title = "Escala"
    
    # Adicionar título
    ws_escala.merge_cells('A1:D1')
    ws_escala['A1'] = f"Escala de Trabalho - {calendar.month_name[mes]} de {ano}"
    ws_escala['A1'].font = Font(bold=True, size=14)
    ws_escala['A1'].alignment = Alignment(horizontal='center')
    
    # Adicionar cabeçalhos
    headers = ["Data", "Dia da Semana", "Turno Matutino", "Turno Vespertino"]
    for col, header in enumerate(headers, start=1):
        cell = ws_escala.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Adicionar dados
    for i, row in df.iterrows():
        ws_escala.cell(row=i+3, column=1, value=row["Data"])
        ws_escala.cell(row=i+3, column=2, value=row["Dia da Semana"])
        ws_escala.cell(row=i+3, column=3, value=row["Turno Matutino"])
        ws_escala.cell(row=i+3, column=4, value=row["Turno Vespertino"])
    
    # Ajustar largura das colunas
    for col in ws_escala.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws_escala.column_dimensions[col[0].column_letter].width = adjusted_width
    
    # Segunda aba - Estatísticas
    ws_stats = workbook.create_sheet(title="Estatísticas")
    
    # Adicionar título
    ws_stats.merge_cells('A1:B1')
    ws_stats['A1'] = "Estatísticas de Alocação"
    ws_stats['A1'].font = Font(bold=True, size=14)
    ws_stats['A1'].alignment = Alignment(horizontal='center')
    
    # Adicionar cabeçalhos
    ws_stats['A2'] = "Funcionário"
    ws_stats['B2'] = "Total de Turnos"
    ws_stats['A2'].font = Font(bold=True)
    ws_stats['B2'].font = Font(bold=True)
    ws_stats['A2'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    ws_stats['B2'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Adicionar dados
    for i, row in estatisticas.iterrows():
        ws_stats.cell(row=i+3, column=1, value=row["Funcionário"])
        ws_stats.cell(row=i+3, column=2, value=row["Total de Turnos"])
    
    # Ajustar largura das colunas
    for col in ws_stats.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws_stats.column_dimensions[col[0].column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    return output

# Interface do Streamlit
st.title("🗓️ Gerador de Escala de Trabalho")

# Criar layout com colunas
col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("Configurações")
    
    # Seleção do mês e ano
    mes = st.selectbox("Selecione o mês", list(range(1, 13)), 
                       index=datetime.today().month - 1, 
                       format_func=lambda x: calendar.month_name[x])
    ano = st.selectbox("Selecione o ano", 
                       list(range(datetime.today().year - 1, datetime.today().year + 5)))
    
    # Opções adicionais
    st.subheader("Opções")
    equilibrar_carga = st.checkbox("Equilibrar carga de trabalho entre funcionários", value=True)
    
    # Inserção da lista de funcionários
    st.subheader("Funcionários")
    funcionarios_input = st.text_area("Digite os nomes dos funcionários (um por linha)")
    funcionarios = [f.strip() for f in funcionarios_input.split("\n") if f.strip()]
    
    # Mostrar número de funcionários e dias úteis
    if funcionarios:
        dias_uteis = sum(1 for dia in range(1, calendar.monthrange(ano, mes)[1] + 1) 
                        if datetime(ano, mes, dia).weekday() < 5)
        st.info(f"Total de funcionários: {len(funcionarios)}")
        st.info(f"Dias úteis no mês: {dias_uteis}")
        st.info(f"Total de turnos a serem preenchidos: {dias_uteis * 2}")

with col2:
    # Interface para exceções
    st.subheader("Exceções")
    excecoes = {}
    
    if funcionarios:
        with st.expander("Configurar exceções"):
            for funcionario in funcionarios:
                st.write(f"**{funcionario}**")
                
                col_data, col_turno = st.columns(2)
                
                with col_data:
                    datas_excecoes = st.text_input(
                        f"Dias que {funcionario} não pode trabalhar (ex: 1, 2, 15)",
                        key=f"data_{funcionario}"
                    )
                
                with col_turno:
                    turnos_excecoes = st.multiselect(
                        f"Turnos indisponíveis",
                        ["Matutino", "Vespertino"],
                        key=f"turno_{funcionario}"
                    )
                
                st.divider()
                
                if datas_excecoes and turnos_excecoes:
                    try:
                        datas_excecoes = [int(d.strip()) for d in datas_excecoes.split(",") if d.strip().isdigit()]
                        for data in datas_excecoes:
                            for turno in turnos_excecoes:
                                excecoes.setdefault((data, turno), []).append(funcionario)
                    except ValueError:
                        st.error(f"Formato inválido para as datas de {funcionario}")

    # Botão para gerar escala
    if st.button("Gerar Escala", type="primary"):
        if not funcionarios:
            st.warning("Por favor, insira pelo menos um funcionário.")
        else:
            with st.spinner("Gerando escala..."):
                escala_df, estatisticas_df = gerar_escala(mes, ano, funcionarios, excecoes, equilibrar_carga)
                
                # Mostrar resultados em tabs
                tab1, tab2, tab3 = st.tabs(["Escala", "Estatísticas", "Visualização"])
                
                with tab1:
                    st.dataframe(escala_df, use_container_width=True)
                    
                    # Opções de download
                    col_csv, col_excel = st.columns(2)
                    with col_csv:
                        csv = escala_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            label="📄 Baixar como CSV",
                            data=csv,
                            file_name=f"escala_{mes}_{ano}.csv",
                            mime="text/csv"
                        )
                    
                    with col_excel:
                        excel_data = exportar_excel(escala_df, estatisticas_df, mes, ano)
                        st.download_button(
                            label="📊 Baixar como Excel",
                            data=excel_data,
                            file_name=f"escala_{mes}_{ano}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                with tab2:
                    st.subheader("Distribuição de turnos por funcionário")
                    st.dataframe(estatisticas_df, use_container_width=True)
                    
                    # Gráfico de barras para visualizar distribuição de turnos
                    chart = alt.Chart(estatisticas_df).mark_bar().encode(
                        x=alt.X('Funcionário', sort='-y'),
                        y=alt.Y('Total de Turnos'),
                        color=alt.Color('Funcionário', legend=None)
                    ).properties(
                        height=300
                    )
                    st.altair_chart(chart, use_container_width=True)
                
                with tab3:
                    st.subheader("Visualização da Escala")
                    
                    # Preparar dados para visualização do calendário
                    cal_data = []
                    for i, row in escala_df.iterrows():
                        data = row['Data']
                        dia = int(data.split('/')[0])
                        cal_data.append({
                            'Dia': dia,
                            'Turno': 'Matutino',
                            'Funcionário': row['Turno Matutino'] or '-'
                        })
                        cal_data.append({
                            'Dia': dia,
                            'Turno': 'Vespertino',
                            'Funcionário': row['Turno Vespertino'] or '-'
                        })
                    
                    cal_df = pd.DataFrame(cal_data)
                    
                    # Criar um heatmap da escala
                    heatmap = alt.Chart(cal_df).mark_rect().encode(
                        x=alt.X('Dia:O', title='Dia do Mês'),
                        y=alt.Y('Turno:N', title=None),
                        color=alt.Color('Funcionário:N', legend=alt.Legend(orient='bottom')),
                        tooltip=['Dia', 'Turno', 'Funcionário']
                    ).properties(
                        title=f"Escala de {calendar.month_name[mes]} de {ano}",
                        width=600
                    )
                    
                    # Adicionar rótulos
                    text = alt.Chart(cal_df).mark_text().encode(
                        x=alt.X('Dia:O'),
                        y=alt.Y('Turno:N'),
                        text='Funcionário',
                        color=alt.condition(
                            alt.datum.Funcionário == '-',
                            alt.value('red'),
                            alt.value('white')
                        )
                    )
                    
                    st.altair_chart(heatmap + text, use_container_width=True)
